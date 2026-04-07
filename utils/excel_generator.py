"""
excel_generator.py - SPPD PTMB Balikpapan
==========================================
Generate file Excel (.xlsx) untuk laporan perjalanan dinas.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

BULAN_ID = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_TOTAL_FILL  = PatternFill("solid", fgColor="F2F2F2")


def _fmt_tgl(tgl_str: str) -> str:
    """'2026-01-04' → '04-Jan-26'"""
    if not tgl_str:
        return ""
    try:
        from datetime import datetime
        d = datetime.strptime(str(tgl_str)[:10], "%Y-%m-%d")
        bln = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Ags","Sep","Okt","Nov","Des"][d.month - 1]
        return f"{d.day:02d}-{bln}-{str(d.year)[2:]}"
    except Exception:
        return str(tgl_str)[:10]


def generate_excel_realisasi(groups: list, bulan: int, tahun: int) -> BytesIO:
    """Generate Excel flat (no merge) untuk laporan realisasi.

    groups: output dari get_sppd_realisasi_laporan() —
            [{"visum": {...}, "sppd_rows": [...]}, ...]
    bulan, tahun: int

    Kolom:
    A:No  B:Tgl Brgkt  C:Tgl Kmbli  D:Uraian  E:Kota  F:No.SPD
    G:Nama  H:Jabatan  I:No.Voucher  J:SPPD  K:Tiket  L:Hotel  M:BiayaLain  N:Total
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Realisasi {BULAN_ID[bulan]} {tahun}"

    # ── Judul
    ws.merge_cells("A1:N1")
    ws["A1"] = "REALISASI SURAT PERMINTAAN PERJALANAN DINAS (SPPD)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:N2")
    ws["A2"] = f"BULAN {BULAN_ID[bulan].upper()} {tahun}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Header kolom (baris 4)
    headers = [
        "No", "Tgl Berangkat", "Tgl Kembali", "Uraian Kegiatan", "Kota", "No. SPD",
        "Nama", "Jabatan", "No. Voucher",
        "SPPD (Rp)", "Tiket Pesawat (Rp)", "Hotel (Rp)", "Biaya Lain-lain (Rp)", "Total (Rp)",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    ws.row_dimensions[4].height = 30

    # ── Data rows
    row_num = 5
    total_j = total_k = total_l = total_m = total_n = 0
    num_format = '#,##0'

    for no_urut, group in enumerate(groups, start=1):
        v = group["visum"]
        tgl_b = _fmt_tgl(v.get("tanggal_berangkat", ""))
        tgl_k = _fmt_tgl(v.get("tanggal_kembali", ""))
        uraian = v.get("keperluan", "") or ""
        kota   = v.get("tujuan", "") or ""
        no_spd = v.get("nomor_spd", "") or ""

        for sppd in group["sppd_rows"]:
            data = [
                no_urut,
                tgl_b,
                tgl_k,
                uraian,
                kota,
                no_spd,
                (sppd.get("nama") or "").title(),
                (sppd.get("jabatan") or "").title(),
                sppd.get("nomor_voucher") or "",
                sppd.get("uang_saku", 0),
                sppd.get("tiket", 0),
                sppd.get("hotel", 0),
                sppd.get("biaya_lain", 0),
                sppd.get("total", 0),
            ]
            for col, val in enumerate(data, start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=(col in (4, 7, 8)))
                if col >= 10:  # kolom angka
                    cell.number_format = num_format

            total_j += sppd.get("uang_saku", 0)
            total_k += sppd.get("tiket", 0)
            total_l += sppd.get("hotel", 0)
            total_m += sppd.get("biaya_lain", 0)
            total_n += sppd.get("total", 0)
            row_num += 1

    # ── Baris TOTAL
    total_row = row_num
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.merge_cells(f"A{total_row}:I{total_row}")
    ws[f"A{total_row}"].alignment = Alignment(horizontal="center")
    ws[f"A{total_row}"].fill = _TOTAL_FILL

    for col, val in zip(range(10, 15), [total_j, total_k, total_l, total_m, total_n]):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = _TOTAL_FILL
        cell.number_format = num_format
        cell.border = _BORDER

    for col in range(1, 10):
        ws.cell(row=total_row, column=col).border = _BORDER

    # ── Lebar kolom
    col_widths = [5, 12, 12, 35, 12, 18, 22, 22, 14, 14, 14, 12, 16, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
